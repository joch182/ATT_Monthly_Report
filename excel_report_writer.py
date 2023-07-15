import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

class ReportWriter():
    def __init__(self) -> None:
        self.wb = Workbook()

    def create_sheet(self, sheet_name):
        self.wb.create_sheet(sheet_name)

    def get_sheet(self, title):
        return self.wb[title]

    @staticmethod
    def save_df_sheet(ws, df):
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)

    def save_excel_report(self, rat):
        self.date = datetime.now().strftime("%Y%m%d")
        self.report_path = "reports_output/"+self.date
        if os.path.isdir(self.report_path):
            pass
        else:
            os.mkdir(self.report_path)
        self.now = datetime.now().strftime("%Y%m%d%H%M%S")
        self.wb.save(self.report_path+"/"+self.now+"_"+rat+"_Monthly_Report.xlsx")
